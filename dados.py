"""
Camada de acesso ao Excel — todas as operações de leitura e escrita.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, threading

EXCEL_PATH = os.environ.get("EXCEL_PATH", "essencia_banco.xlsx")
_lock = threading.Lock()

# ── Helpers de estilo ──────────────────────────────────────────────────────────
def hfill(hex_): return PatternFill("solid", start_color=hex_, fgColor=hex_)
def thin():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

COR_NIVEL = {
    "N0": "F5C6C6",
    "N1": "E6F1FB",
    "N2": "EAF3DE",
    "N3": "FAEEDA",
    "N4": "E1F5EE",
}

NIVEIS = [
    "N1 - Ideia",
    "N2 - Planejamento",
    "N3 - Execução",
    "N4 - Implementado",
    "N0 - Cancelada",
]

MESES = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]

COLS_OP = [
    "ID","Nível","Descrição","Comentário da Semana",
    "Grupo Contábil","Conta Orçamento","Conta Contábil",
    "Dono da Oportunidade","CC Dono",
    "Est. Jan/26","Est. Fev/26","Est. Mar/26","Est. Abr/26",
    "Est. Mai/26","Est. Jun/26","Est. Jul/26","Est. Ago/26",
    "Est. Set/26","Est. Out/26","Est. Nov/26","Est. Dez/26",
    "Total 2026","Total 2027","Total 2028",
    "Craque","Filial","Área","Frente de Negócio",
    "Data Cadastro (N1)",
    "Data Esperada N3","Data Esperada N4",
    "Data Realizada N0","Data Realizada N2",
    "Data Realizada N3","Data Realizada N4",
    "Histórico de Níveis",
]

# ── Leitura ────────────────────────────────────────────────────────────────────
def ler_oportunidades() -> pd.DataFrame:
    with _lock:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Oportunidades", dtype=str)
        df.columns = COLS_OP
        df = df.fillna("")
        # Converte colunas de valor para numérico
        cols_val = [c for c in COLS_OP if c.startswith("Est.") or c.startswith("Total")]
        for col in cols_val:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df

def ler_usuarios() -> pd.DataFrame:
    with _lock:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Usuários", dtype=str)
        df.columns = ["Login","Nome","Perfil","Ativo"]
        return df[df["Ativo"].str.lower() == "sim"].fillna("")

def ler_craques() -> pd.DataFrame:
    with _lock:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Craques", dtype=str)
        df.columns = ["Login","Nome","Filial","Área","Frente"]
        return df.fillna("")

def ler_lideres() -> pd.DataFrame:
    with _lock:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Líderes", dtype=str)
        df.columns = ["Login","Nome","Frente"]
        return df.fillna("")

def ler_plano_contas() -> pd.DataFrame:
    with _lock:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Plano de Contas", dtype=str)
        df.columns = ["Grupo","ContaOrc","ContaCont","Código"]
        return df.fillna("")

def ler_orcado() -> pd.DataFrame:
    with _lock:
        cols = ["Frente","Ano"] + MESES + ["Total"]
        df = pd.read_excel(EXCEL_PATH, sheet_name="Orçado", dtype=str)
        df.columns = cols
        for col in MESES + ["Total"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df

# ── Auth ───────────────────────────────────────────────────────────────────────
def autenticar(login: str):
    """Retorna dict com dados do usuário ou None se não encontrado."""
    usuarios = ler_usuarios()
    row = usuarios[usuarios["Login"].str.lower() == login.lower()]
    if row.empty:
        return None
    u = row.iloc[0]
    dados = {"login": u["Login"], "nome": u["Nome"], "perfil": u["Perfil"]}
    if u["Perfil"] == "craque":
        craques = ler_craques()
        cq = craques[craques["Login"].str.lower() == login.lower()]
        if not cq.empty:
            dados["filial"] = cq.iloc[0]["Filial"]
            dados["area"]   = cq.iloc[0]["Área"]
            dados["frente"] = cq.iloc[0]["Frente"]
    elif u["Perfil"] == "lider":
        lideres = ler_lideres()
        lid = lideres[lideres["Login"].str.lower() == login.lower()]
        if not lid.empty:
            dados["frente"] = lid.iloc[0]["Frente"]
    return dados

# ── Próximo ID ─────────────────────────────────────────────────────────────────
def proximo_id(df: pd.DataFrame) -> str:
    if df.empty or df["ID"].eq("").all():
        return "001"
    ids = df["ID"].str.extract(r"(\d+)")[0].dropna().astype(int)
    return str(ids.max() + 1).zfill(3)

# ── Salvar linha no Excel ──────────────────────────────────────────────────────
def _reescrever_oportunidades(df: pd.DataFrame):
    """Reescreve a aba Oportunidades preservando as demais abas."""
    with _lock:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Oportunidades"]

        # Limpa dados antigos (mantém cabeçalho linha 1)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fill_type=None)

        for ri, (_, row) in enumerate(df.iterrows(), 2):
            nivel_key = str(row.get("Nível","N1"))[:2]
            bg = COR_NIVEL.get(nivel_key, "FFFFFF")
            for ci, col in enumerate(COLS_OP, 1):
                val = row.get(col, "")
                if pd.isna(val): val = ""
                c = ws.cell(ri, ci, val)
                c.font   = Font(name="Arial", size=10)
                c.fill   = hfill(bg)
                c.border = thin()
                c.alignment = Alignment(
                    horizontal="center" if ci in [1,2,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,29,30,31,32,33,34,35]
                    else "left", vertical="center", wrap_text=True
                )
                if ci in range(10, 25):
                    c.number_format = 'R$ #,##0'
            ws.row_dimensions[ri].height = 22

        wb.save(EXCEL_PATH)

# ── Cadastrar oportunidade ─────────────────────────────────────────────────────
def cadastrar_oportunidade(dados: dict, usuario: dict) -> str:
    df = ler_oportunidades()
    novo_id = proximo_id(df)
    hoje = datetime.now().strftime("%d/%m/%Y")

    nova = {col: "" for col in COLS_OP}
    nova.update({
        "ID":                    novo_id,
        "Nível":                 "N1 - Ideia",
        "Descrição":             dados.get("descricao",""),
        "Comentário da Semana":  "",
        "Grupo Contábil":        dados.get("grupo",""),
        "Conta Orçamento":       dados.get("conta_orc",""),
        "Conta Contábil":        dados.get("conta_cont",""),
        "Dono da Oportunidade":  dados.get("dono",""),
        "CC Dono":               dados.get("cc_dono",""),
        "Total 2027":            dados.get("total_2027", 0),
        "Total 2028":            dados.get("total_2028", 0),
        "Craque":                usuario.get("nome",""),
        "Filial":                usuario.get("filial",""),
        "Área":                  usuario.get("area",""),
        "Frente de Negócio":     usuario.get("frente",""),
        "Data Cadastro (N1)":    hoje,
        "Histórico de Níveis":   f"N1 - Ideia ({hoje})",
    })
    # Estimativas mensais
    for m in MESES:
        nova[f"Est. {m}/26"] = dados.get(f"est_{m.lower()}", 0)
    # Total 2026 automático
    nova["Total 2026"] = sum(dados.get(f"est_{m.lower()}", 0) for m in MESES)

    df = pd.concat([df, pd.DataFrame([nova])], ignore_index=True)
    _reescrever_oportunidades(df)
    return novo_id

# ── Atualizar campos ───────────────────────────────────────────────────────────
def atualizar_oportunidade(id_: str, campos: dict, usuario: dict):
    df = ler_oportunidades()
    idx = df.index[df["ID"] == id_]
    if idx.empty:
        raise ValueError("Oportunidade não encontrada")
    i = idx[0]
    for k, v in campos.items():
        if k in df.columns:
            df.at[i, k] = v
    # Recalcula Total 2026
    meses_vals = [df.at[i, f"Est. {m}/26"] for m in MESES]
    df.at[i, "Total 2026"] = sum(float(v) if v != "" else 0 for v in meses_vals)
    _reescrever_oportunidades(df)

# ── Movimentar nível ───────────────────────────────────────────────────────────
def movimentar_nivel(id_: str, novo_nivel: str, usuario: dict):
    df = ler_oportunidades()
    idx = df.index[df["ID"] == id_]
    if idx.empty:
        raise ValueError("Oportunidade não encontrada")
    i = idx[0]
    hoje = datetime.now().strftime("%d/%m/%Y")
    nivel_atual = df.at[i, "Nível"]

    # Valida permissão de movimentação
    perfil = usuario.get("perfil","")
    erros = []
    if novo_nivel == "N0 - Cancelada":
        if perfil not in ("lider","adm"):
            erros.append("Apenas Líder da Frente ou Adm podem cancelar.")
    elif novo_nivel == "N2 - Planejamento":
        if perfil != "adm":
            erros.append("Apenas Adm podem mover para N2.")
    elif novo_nivel == "N3 - Execução":
        if perfil not in ("lider","adm"):
            erros.append("Apenas Líder da Frente ou Adm podem mover para N3.")
    elif novo_nivel == "N4 - Implementado":
        if perfil != "adm":
            erros.append("Apenas Adm podem mover para N4.")
    if erros:
        raise PermissionError(erros[0])

    # Atualiza datas realizadas
    if novo_nivel == "N0 - Cancelada":
        df.at[i, "Data Realizada N0"] = hoje
    elif novo_nivel == "N2 - Planejamento":
        df.at[i, "Data Realizada N2"] = hoje
    elif novo_nivel == "N3 - Execução":
        df.at[i, "Data Realizada N3"] = hoje
    elif novo_nivel == "N4 - Implementado":
        df.at[i, "Data Realizada N4"] = hoje

    # Histórico
    hist = df.at[i, "Histórico de Níveis"]
    df.at[i, "Histórico de Níveis"] = hist + f" → {novo_nivel} ({hoje})"
    df.at[i, "Nível"] = novo_nivel
    _reescrever_oportunidades(df)

# ── Adicionar comentário ───────────────────────────────────────────────────────
def adicionar_comentario(id_: str, texto: str, usuario: dict):
    df = ler_oportunidades()
    idx = df.index[df["ID"] == id_]
    if idx.empty:
        raise ValueError("Oportunidade não encontrada")
    i = idx[0]
    hoje = datetime.now().strftime("%d/%m/%Y %H:%M")
    autor = usuario.get("nome","?")
    comentario_atual = df.at[i, "Comentário da Semana"]
    novo = f"[{hoje} — {autor}] {texto}"
    df.at[i, "Comentário da Semana"] = (comentario_atual + "\n" + novo).strip()
    _reescrever_oportunidades(df)
