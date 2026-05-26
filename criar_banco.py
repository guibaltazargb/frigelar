"""
Gerador do banco de dados Excel do Programa Essência.
Execute uma vez para criar o arquivo essencia_banco.xlsx.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from datetime import datetime

EXCEL_PATH = "essencia_banco.xlsx"

AZ1, AZ2, AZ3 = "0D2240", "185FA5", "B5D4F4"
VD1, VD2      = "1B3D06", "3B6D11"

def hfill(hex_): return PatternFill("solid", start_color=hex_, fgColor=hex_)
def hfont(hex_, bold=False, sz=10): return Font(name="Arial", size=sz, bold=bold, color=hex_)
def thin():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def header_row(ws, cols, row=1, fill=AZ1):
    for ci, txt in enumerate(cols, 1):
        c = ws.cell(row, ci, txt)
        c.font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill  = hfill(fill)
        c.alignment = center()
        c.border = thin()
    ws.row_dimensions[row].height = 28

def criar_excel():
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════
    # ABA: Oportunidades
    # ════════════════════════════════════════════
    ws_op = wb.active
    ws_op.title = "Oportunidades"
    ws_op.sheet_view.showGridLines = False
    ws_op.freeze_panes = "A2"

    cols_op = [
        "ID", "Nível", "Descrição", "Comentário da Semana",
        "Grupo Contábil", "Conta Orçamento", "Conta Contábil",
        "Dono da Oportunidade", "CC Dono",
        "Est. Jan/26", "Est. Fev/26", "Est. Mar/26", "Est. Abr/26",
        "Est. Mai/26", "Est. Jun/26", "Est. Jul/26", "Est. Ago/26",
        "Est. Set/26", "Est. Out/26", "Est. Nov/26", "Est. Dez/26",
        "Total 2026", "Total 2027", "Total 2028",
        "Craque", "Filial", "Área", "Frente de Negócio",
        "Data Cadastro (N1)",
        "Data Esperada N3", "Data Esperada N4",
        "Data Realizada N0", "Data Realizada N2",
        "Data Realizada N3", "Data Realizada N4",
        "Histórico de Níveis",
    ]
    header_row(ws_op, cols_op)

    larguras_op = {
        1:6, 2:18, 3:40, 4:35,
        5:20, 6:22, 7:22,
        8:22, 9:14,
        10:11,11:11,12:11,13:11,14:11,15:11,16:11,17:11,18:11,19:11,20:11,21:11,
        22:12, 23:12, 24:12,
        25:18, 26:14, 27:18, 28:18,
        29:14, 30:14, 31:14,
        32:14, 33:14, 34:14, 35:14,
        36:35,
    }
    for ci, w in larguras_op.items():
        ws_op.column_dimensions[get_column_letter(ci)].width = w

    # Dados de exemplo
    hoje = datetime.now().strftime("%d/%m/%Y")
    exemplos = [
        ("001","N1 - Ideia","Redução consumo energia no CD","",
         "Custos Operacionais","Energia Elétrica","3.1.2.001",
         "João Pereira","CC-001",
         0,0,0,0,0,0,18000,18000,18000,18000,18000,18000,
         90000,216000,216000,
         "Craque Teste","Filial SP","Logística","Operações",
         hoje,"","","","","","","N1 - Ideia"),

        ("002","N2 - Planejamento","Renegociação contrato embalagens","Reunião com fornecedor agendada para semana que vem.",
         "Custos de Materiais","Embalagens","3.2.1.005",
         "Maria Costa","CC-002",
         0,0,0,0,0,42000,42000,42000,42000,42000,42000,42000,
         294000,504000,504000,
         "Carlos Silva","Filial RJ","Compras","Supply",
         hoje,"30/06/2026","31/07/2026","",hoje,"","","N1 - Ideia → N2 - Planejamento"),

        ("003","N3 - Execução","Automação relatórios controladoria","Mapeamento concluído, iniciando desenvolvimento.",
         "Despesas Administrativas","TI e Sistemas","4.1.3.002",
         "Ana Lima","CC-003",
         0,0,0,0,6000,6000,6000,6000,6000,6000,6000,6000,
         54000,72000,72000,
         "Craque Teste","Filial SP","Controladoria","Financeiro",
         hoje,"15/06/2026","30/07/2026","",hoje,hoje,"","N1 - Ideia → N2 - Planejamento → N3 - Execução"),

        ("004","N4 - Implementado","Substituição insumo importado","Implementado com sucesso. Economia confirmada.",
         "Custos de Produção","Matéria Prima","3.3.1.010",
         "Roberto Melo","CC-004",
         47400,47400,47400,47400,47400,47400,47400,47400,47400,47400,47400,47400,
         568800,568800,568800,
         "Ana Lima","Filial MG","Produção","Operações",
         hoje,"01/05/2026","15/05/2026","",hoje,hoje,hoje,
         "N1 - Ideia → N2 - Planejamento → N3 - Execução → N4 - Implementado"),
    ]

    COR_NIVEL = {"N0":"F5C6C6","N1":"E6F1FB","N2":"EAF3DE","N3":"FAEEDA","N4":"E1F5EE"}

    for ri, ex in enumerate(exemplos, 2):
        nivel_key = ex[1][:2]
        bg = COR_NIVEL.get(nivel_key, "FFFFFF")
        for ci, val in enumerate(ex, 1):
            c = ws_op.cell(ri, ci, val)
            c.font   = Font(name="Arial", size=10)
            c.fill   = hfill(bg)
            c.border = thin()
            c.alignment = center() if ci in [1,2,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,29,30,31,32,33,34,35] else left()
            if ci in range(10, 25):
                c.number_format = 'R$ #,##0'
        ws_op.row_dimensions[ri].height = 22

    # ════════════════════════════════════════════
    # ABA: Usuários
    # ════════════════════════════════════════════
    ws_usr = wb.create_sheet("Usuários")
    ws_usr.sheet_view.showGridLines = False
    header_row(ws_usr, ["Login","Nome Completo","Perfil","Ativo"], fill=AZ1)
    usuarios = [
        ("craque_teste",  "Craque Teste",    "craque",        "Sim"),
        ("craque_02",     "Craque Dois",     "craque",        "Sim"),
        ("lider_op",      "Líder Operações", "lider",         "Sim"),
        ("lider_supply",  "Líder Supply",    "lider",         "Sim"),
        ("lider_fin",     "Líder Financeiro","lider",         "Sim"),
        ("controladoria", "Controladoria",   "adm",           "Sim"),
        ("silva",         "Silva",           "adm",           "Sim"),
    ]
    for ri, u in enumerate(usuarios, 2):
        for ci, val in enumerate(u, 1):
            c = ws_usr.cell(ri, ci, val)
            c.font = Font(name="Arial", size=10)
            c.fill = hfill("EAF3DE" if ri % 2 == 0 else "FFFFFF")
            c.border = thin(); c.alignment = left()
        ws_usr.row_dimensions[ri].height = 20
    for col, w in {1:18,2:24,3:14,4:8}.items():
        ws_usr.column_dimensions[get_column_letter(col)].width = w

    # ════════════════════════════════════════════
    # ABA: Craques
    # ════════════════════════════════════════════
    ws_cq = wb.create_sheet("Craques")
    ws_cq.sheet_view.showGridLines = False
    header_row(ws_cq, ["Login","Nome Completo","Filial","Área","Frente de Negócio"], fill=AZ1)
    craques = [
        ("craque_teste","Craque Teste","Filial SP","Logística","Operações"),
        ("craque_02",   "Craque Dois", "Filial RJ","Compras",  "Supply"),
    ]
    for ri, cq in enumerate(craques, 2):
        for ci, val in enumerate(cq, 1):
            c = ws_cq.cell(ri, ci, val)
            c.font = Font(name="Arial", size=10)
            c.fill = hfill("E6F1FB" if ri % 2 == 0 else "FFFFFF")
            c.border = thin(); c.alignment = left()
        ws_cq.row_dimensions[ri].height = 20
    for col, w in {1:18,2:24,3:16,4:18,5:20}.items():
        ws_cq.column_dimensions[get_column_letter(col)].width = w

    # ════════════════════════════════════════════
    # ABA: Líderes
    # ════════════════════════════════════════════
    ws_lid = wb.create_sheet("Líderes")
    ws_lid.sheet_view.showGridLines = False
    header_row(ws_lid, ["Login","Nome Completo","Frente de Negócio"], fill=AZ1)
    lideres = [
        ("lider_op",     "Líder Operações", "Operações"),
        ("lider_supply",  "Líder Supply",   "Supply"),
        ("lider_fin",     "Líder Financeiro","Financeiro"),
    ]
    for ri, l in enumerate(lideres, 2):
        for ci, val in enumerate(l, 1):
            c = ws_lid.cell(ri, ci, val)
            c.font = Font(name="Arial", size=10)
            c.fill = hfill("FAEEDA" if ri % 2 == 0 else "FFFFFF")
            c.border = thin(); c.alignment = left()
        ws_lid.row_dimensions[ri].height = 20
    for col, w in {1:18,2:24,3:20}.items():
        ws_lid.column_dimensions[get_column_letter(col)].width = w

    # ════════════════════════════════════════════
    # ABA: Plano de Contas
    # ════════════════════════════════════════════
    ws_pc = wb.create_sheet("Plano de Contas")
    ws_pc.sheet_view.showGridLines = False
    header_row(ws_pc, ["Grupo Contábil","Conta Orçamento","Conta Contábil","Código"], fill=AZ1)
    planos = [
        ("Custos Operacionais",    "Energia Elétrica",  "Energia Elétrica CD",     "3.1.2.001"),
        ("Custos Operacionais",    "Energia Elétrica",  "Energia Elétrica Admin",  "3.1.2.002"),
        ("Custos Operacionais",    "Fretes e Logística","Frete Saída",             "3.1.3.001"),
        ("Custos Operacionais",    "Fretes e Logística","Frete Retorno",           "3.1.3.002"),
        ("Custos de Materiais",    "Embalagens",        "Embalagem Primária",      "3.2.1.005"),
        ("Custos de Materiais",    "Embalagens",        "Embalagem Secundária",    "3.2.1.006"),
        ("Custos de Materiais",    "Matéria Prima",     "Insumo Nacional",         "3.2.2.001"),
        ("Custos de Produção",     "Matéria Prima",     "Insumo Importado",        "3.3.1.010"),
        ("Custos de Produção",     "Manutenção",        "Manutenção Preventiva",   "3.3.2.001"),
        ("Despesas Administrativas","TI e Sistemas",    "Licenças de Software",    "4.1.3.001"),
        ("Despesas Administrativas","TI e Sistemas",    "Desenvolvimento",         "4.1.3.002"),
        ("Despesas Administrativas","Serviços Terceiros","Consultoria",            "4.1.4.001"),
    ]
    for ri, p in enumerate(planos, 2):
        for ci, val in enumerate(p, 1):
            c = ws_pc.cell(ri, ci, val)
            c.font = Font(name="Arial", size=10)
            c.fill = hfill("E6F1FB" if ri % 2 == 0 else "FFFFFF")
            c.border = thin(); c.alignment = left()
        ws_pc.row_dimensions[ri].height = 20
    for col, w in {1:26,2:24,3:28,4:14}.items():
        ws_pc.column_dimensions[get_column_letter(col)].width = w

    # ════════════════════════════════════════════
    # ABA: Orçado
    # ════════════════════════════════════════════
    ws_orc = wb.create_sheet("Orçado")
    ws_orc.sheet_view.showGridLines = False
    meses = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    header_row(ws_orc, ["Frente de Negócio","Ano"] + meses + ["Total"], fill=VD1)
    frentes_orc = ["Operações","Supply","Financeiro","Corporativo"]
    for ri, (frente, ano) in enumerate([(f,a) for f in frentes_orc for a in [2026,2027,2028]], 2):
        ws_orc.cell(ri, 1, frente).font = Font(name="Arial", size=10)
        ws_orc.cell(ri, 1).fill = hfill("EAF3DE" if ri % 2 == 0 else "FFFFFF")
        ws_orc.cell(ri, 1).border = thin()
        ws_orc.cell(ri, 2, ano).font = Font(name="Arial", size=10)
        ws_orc.cell(ri, 2).fill = hfill("EAF3DE" if ri % 2 == 0 else "FFFFFF")
        ws_orc.cell(ri, 2).border = thin()
        ws_orc.cell(ri, 2).alignment = center()
        for ci in range(3, 15):
            c = ws_orc.cell(ri, ci, 0)
            c.font = Font(name="Arial", size=10)
            c.fill = hfill("EAF3DE" if ri % 2 == 0 else "FFFFFF")
            c.border = thin(); c.alignment = center()
            c.number_format = 'R$ #,##0'
        # Total
        last_col = get_column_letter(14)
        formula = f"=SUM(C{ri}:{last_col}{ri})"
        c_tot = ws_orc.cell(ri, 15, formula)
        c_tot.font = Font(name="Arial", size=10, bold=True)
        c_tot.fill = hfill("C0DD97")
        c_tot.border = thin(); c_tot.alignment = center()
        c_tot.number_format = 'R$ #,##0'
        ws_orc.row_dimensions[ri].height = 20
    for col, w in {1:20,2:8,3:10,4:10,5:10,6:10,7:10,8:10,9:10,10:10,11:10,12:10,13:10,14:10,15:12}.items():
        ws_orc.column_dimensions[get_column_letter(col)].width = w

    wb.save(EXCEL_PATH)
    print(f"✅ Excel criado: {EXCEL_PATH}")
    print(f"   Abas: {[ws.title for ws in wb.worksheets]}")

if __name__ == "__main__":
    criar_excel()
